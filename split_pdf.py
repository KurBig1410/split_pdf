import os
import openpyxl
# from openpyxl.utils import get_column_letter
from PyPDF2 import PdfReader, PdfWriter
import re  # Для очистки имени файла
import urllib.parse

def get_merged_cell_value(sheet, cell):
    """Получает значение объединённой ячейки, если она есть"""
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            return sheet.cell(row=min_row, column=min_col).value
    return cell.value

def clean_int(value):
    """Преобразует значение в целое число, убирая текст"""
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        digits = ''.join(filter(str.isdigit, value))  # Удаляем всё, кроме цифр
        return int(digits) if digits else None
    return None



def sanitize_filename(filename):
    """Очищает имя файла от запрещённых символов, ограничивает длину и декодирует URL-encoded символы"""
    
    # Декодируем URL-кодированные символы
    filename = urllib.parse.unquote(filename)
    
    # Убираем спецсимволы
    filename = re.sub(r'[\\/*?:""<>|;]', "", filename)
    
    # Заменяем пробелы на "_"
    filename = filename.replace(" ", "_")
    
    # Разделяем строку на слова
    words = filename.split("_")
    
    # Если количество слов больше 15, оставляем только первые 15
    if len(words) > 5:
        filename = "_".join(words[:5])
    
    return filename
    
    return filename

def split_pdf(input_pdf, input_excel, output_folder):
    # Открываем Excel-файл
    workbook = openpyxl.load_workbook(input_excel)
    sheet = workbook.active

    # Открываем PDF-файл
    pdf_reader = PdfReader(input_pdf)
    total_pages = len(pdf_reader.pages)

    # Создаём папку для сохранения файлов, если её нет
    os.makedirs(output_folder, exist_ok=True)

    current_page = 0  # Начинаем с первой страницы

    for row in sheet.iter_rows(min_row=17, max_col=sheet.max_column, values_only=False):
        row_num = row[2].row  # Номер строки в Excel
        doc_number = get_merged_cell_value(sheet, row[2])  # H - Номер документа
        org_name = get_merged_cell_value(sheet, row[3])  # K - Организация
        page_count = clean_int(get_merged_cell_value(sheet, row[11]))  # L - Количество листов
        
        if page_count is None or page_count <= 0:
            # print(f"Ошибка: некорректное значение страниц в строке {row_num} ({get_merged_cell_value(sheet, row[11])})")
            continue
        page_count *= 2

        # Безопасное имя файла
        doc_number = doc_number if doc_number else "б/н"
        org_name = org_name if org_name else "неизвестная_организация"
        safe_filename = f"{doc_number}_{sanitize_filename(org_name)}.pdf"

        output_pdf_path = os.path.join(output_folder, safe_filename)
        hyper_path = f"Скан/{safe_filename}"

        # Проверяем, не выходит ли за пределы PDF
        if current_page + page_count > total_pages:
            print(f"Ошибка: превышено количество страниц в строке {row_num}")
            break

        # Разделяем PDF
        pdf_writer = PdfWriter()
        for _ in range(page_count):
            pdf_writer.add_page(pdf_reader.pages[current_page])
            current_page += 1

        # Сохраняем отдельный PDF
        with open(output_pdf_path, "wb") as output_pdf:
            pdf_writer.write(output_pdf)

        # Добавляем гиперссылку в колонку Q (17-я колонка)
        hyperlink_cell = sheet[f"Q{row_num}"]
        hyperlink_cell.hyperlink = hyper_path
        hyperlink_cell.value = f"{doc_number}_{org_name}"
        hyperlink_cell.style = "Hyperlink"

    # Сохраняем обновленный реестр
    workbook.save(input_excel)
    workbook.close()
    print("Разделение завершено и ссылки добавлены!")

# Пример вызова:
# input_pdf = "TEST/Книга 1.2.2.pdf"
# input_excel = "TEST/Реестр книга 1.2.2.xlsx"
# output_folder = "TEST/Скан"
# split_pdf(input_pdf, input_excel, output_folder)